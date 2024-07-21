import argparse
import uuid

import openpyxl

"""
Expense manager format
Date: datetime
Category: str
Income: float
Expense: float
Description: str

Money manager format
Date: mm/dd/yyyy
Account: str (Main)
Category: str
Subcategory: str
Note: str
Amount: float
Income/Expense: str (Income/Expense)
Description:
"""


def expense_to_money():
    parser = argparse.ArgumentParser(description='Expense manager to money manager')
    parser.add_argument('filename', type=str, help='File name to parse')
    args = parser.parse_args()
    filename = args.filename
    expense_manager_workbook = openpyxl.load_workbook(filename)
    expense_manager_sheet = expense_manager_workbook.active
    money_manager_workbook = openpyxl.Workbook()
    money_manager_sheet = money_manager_workbook.active

    count = 1
    money_manager_sheet.append(
        ['Date', 'Account', 'Category', 'Subcategory', 'Note', 'Amount', 'Income/Expense', 'Description']
    )

    for row in expense_manager_sheet.iter_rows(values_only=True):
        expense_date = row[0].strftime('%m/%d/%Y')
        category = row[1] or 'Other'
        income = row[2]
        expense = row[3]
        description = row[4]

        income_or_expense = 'Income' if income else 'Expense'
        amount = income if income else expense

        new_row = [expense_date, 'Main', category, '', '', amount, income_or_expense, description]

        money_manager_sheet.append(new_row)

    for row in money_manager_sheet.iter_rows(values_only=True):
        count += 1
        if count % 100 == 0:
            print(row)
    money_manager_workbook.save(f'money_manager_{uuid.uuid4()}.xlsx')


if __name__ == '__main__':
    expense_to_money()
